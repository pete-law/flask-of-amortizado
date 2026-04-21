import anthropic
import json
import openpyxl
import os
from dotenv import load_dotenv
from pathlib import Path
load_dotenv(Path.home() / "secrets" / ".env")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from edgar import Company, set_identity
from datetime import datetime

# Setup
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY2"))
set_identity(os.getenv("EDGAR_EMAIL"))

CURRENT_YEAR = datetime.now().year
MATURITY_WINDOW = 2

def get_filing_text(company, form_type, limit=3):
    try:
        filings = company.get_filings(form=form_type)
        if not filings:
            return f"No {form_type} filings found", None
        
        text = ""
        url = None
        for i in range(min(limit, len(filings))):
            filing = filings[i]
            text += str(filing.text())[:5000]
            if url is None:
                try:
                    url = filing.document.url
                except:
                    try:
                        url = filing.url
                    except:
                        url = None
        return text, url
    except Exception as e:
        return f"Could not retrieve {form_type}: {str(e)}", None
    
def analyze_company(ticker):
    print(f"\nProcessing {ticker}...")

    company = Company(ticker)

    print(f"  Fetching 10-K...")
    tenk_text, tenk_url = get_filing_text(company, "10-K")

    print(f"  Fetching 8-K...")
    eightk_text, eightk_url = get_filing_text(company, "8-K")

    print(f"  Fetching S-1...")
    s1_text, s1_url = get_filing_text(company, "S-1")

    filing_urls = {
        "10-K": tenk_url,
        "8-K": eightk_url,
        "S-1": s1_url
    }

    print(f"  Fetching balance sheet...")
    try:
        balance_sheet = str(company.get_financials().balance_sheet())
    except:
        balance_sheet = "Balance sheet unavailable"

    prompt = f"""
You are a legal and financial analyst. Analyze these SEC filings for {ticker} and return ONLY a JSON object with no other text, no markdown, no backticks.

Today's year is {CURRENT_YEAR}. Flag any debt maturing within {MATURITY_WINDOW} years as upcoming.

Express all amounts in the most natural denomination:
- Use thousands for amounts under $1 million (e.g. "$500 thousand")
- Use millions for $1 million to $999 million (e.g. "$500 million")
- Use billions for $1 billion and above (e.g. "$5.25 billion")

Return exactly this structure:
{{
    "ticker": "{ticker}",
    "company_name": "full company name",
    "total_debt": "e.g. $98,657 million",
    "interest_rate_range": "e.g. 0% - 3.6%",
    "flag": "Describe any data quality issues here and be watchful for debt double counting (e.g. both short term and upcoming long term debt listed both together and separately), or leave as empty string if none",
    "notes": "Genuine observations only, no flags here. Keep under 200 characters.",
    Specifically check: if both 'short-term debt' and 'current portion of long-term debt' appear as separate line items, flag this as potential double counting in total debt calculation. 
    Then on a new line add any other genuine observations. Leave empty if nothing notable.",
    "debt_items": [
        {{
            "debt_type": "e.g. Commercial Paper / Term Loan / Senior Notes",
            "amount": "e.g. $7,979 million",
            "interest_rate": "e.g. 1.625%",
            "maturity_year": "e.g. 2026 or Not disclosed",
            "is_upcoming": true or false,
            "borrower_counsel": "law firm name or Not disclosed",
            "lender_counsel": "law firm name or Not disclosed"
            "source_filing": "10-K or 8-K or S-1 — whichever filing this instrument was found in",
        }}
    ]
}}

BALANCE SHEET DATA:
{balance_sheet}

10-K FILING:
{tenk_text[:3000]}

8-K FILING:
{eightk_text[:3000]}

S-1 FILING:
{s1_text[:3000]}
"""

    print(f"  Analyzing with Claude...")
    message = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}]
    )

    try:
        text = message.content[0].text
        # Strip any markdown backticks Claude might have added
        text = text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        result = json.loads(text.strip())
        result["filing_urls"] = filing_urls
        return result
    except:
        return {
            "ticker": ticker,
            "company_name": "Parse error",
            "flag": "Claude returned malformed JSON for this ticker",
            "notes": message.content[0].text[:200],
            "total_debt": "Error",
            "interest_rate_range": "Error",
            "debt_items": []
        }

def create_excel(results, filename="sec_analysis.xlsx"):
    wb = openpyxl.Workbook()

    # ── SUMMARY SHEET ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    upcoming_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    summary_headers = [
        "Ticker", "Company Name", "Total Debt",
        "# Instruments", "Interest Rate Range",
        "Upcoming Maturities (2yr)", "Notes"
    ]

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, result in enumerate(results, 2):
        debt_items = result.get("debt_items", [])
        upcoming = [
            f"{item.get('amount', 'N/A')} {item.get('debt_type', '')} ({item.get('maturity_year', '')})"
            for item in debt_items if item.get("is_upcoming")
        ]
        upcoming_text = "\n".join(upcoming) if upcoming else "None within 2 years"
        num_instruments = len(debt_items)

        ws_summary.cell(row=row, column=1, value=result.get("ticker", "N/A")).alignment = Alignment(vertical="center", horizontal="center")
        ws_summary.cell(row=row, column=2, value=result.get("company_name", "N/A")).alignment = Alignment(vertical="center")
        ws_summary.cell(row=row, column=3, value=result.get("total_debt", "N/A")).alignment = Alignment(vertical="center", horizontal="center")
        ws_summary.cell(row=row, column=4, value=num_instruments).alignment = Alignment(vertical="center", horizontal="center")
        ws_summary.cell(row=row, column=5, value=result.get("interest_rate_range", "N/A")).alignment = Alignment(vertical="center", horizontal="center")

        upcoming_lines = [
            f"• {item.get('amount', 'N/A')} {item.get('debt_type', '')} ({item.get('maturity_year', '')})"
            for item in debt_items if item.get("is_upcoming")
        ]
        upcoming_text = "\n".join(upcoming_lines) if upcoming_lines else "None within 2 years"

        upcoming_cell = ws_summary.cell(row=row, column=6, value=upcoming_text)
        upcoming_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        if upcoming and upcoming_text != "None within 2 years":
            upcoming_cell.fill = upcoming_fill

        notes_text = result.get("notes", "").replace("\n\n", "\n").strip()
        flag_text = result.get("flag", "").strip()
        notes_raw = result.get("notes", "").replace("\n\n", "\n").strip()

        if flag_text:
            notes_full = f"*** FLAG: {flag_text}\n\n{notes_raw}".strip()
        else:
            notes_full = notes_raw

        notes_short = (notes_full[:120] + "...") if len(notes_full) > 120 else notes_full
        notes_cell = ws_summary.cell(row=row, column=7, value=notes_short)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="center")

        if notes_full:
            comment = Comment(notes_full, "SEC Analysis Tool")
            comment.width = 300
            comment.height = 200
            notes_cell.comment = comment
        for col in range(1, len(summary_headers) + 1):
            ws_summary.cell(row=row, column=col).border = Border(bottom=Side(style="medium", color="1F4E79"))
    
    # Column widths for summary
    summary_widths = [10, 25, 15, 12, 20, 45, 40]
    for col, width in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = width

    ws_summary.row_dimensions[1].height = 20
    for row_idx, result in enumerate(results, 2):
        upcoming = [i for i in result.get("debt_items", []) if i.get("is_upcoming")]
        height = max(60, 40 + (len(upcoming) * 20))
        ws_summary.row_dimensions[row_idx].height = height    
    # ── DETAIL SHEET ───────────────────────────────────────────
    ws_detail = wb.create_sheet(title="Detail")

    detail_headers = [
        "Ticker", "Company Name", "Debt Type", "Amount",
        "Interest Rate", "Maturity Year",
        "Borrower Counsel", "Lender Counsel", "Source Filing"
    ]

    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    detail_row = 2
    for result in results:
        ticker = result.get("ticker", "N/A")
        company_name = result.get("company_name", "N/A")
        for item in result.get("debt_items", []):
            ws_detail.cell(row=detail_row, column=1, value=ticker)
            ws_detail.cell(row=detail_row, column=2, value=company_name)
            ws_detail.cell(row=detail_row, column=3, value=item.get("debt_type", "N/A"))
            ws_detail.cell(row=detail_row, column=4, value=item.get("amount", "N/A"))
            ws_detail.cell(row=detail_row, column=5, value=item.get("interest_rate", "N/A"))

            maturity = item.get("maturity_year", "N/A")
            maturity_cell = ws_detail.cell(row=detail_row, column=6, value=maturity)
            if item.get("is_upcoming"):
                maturity_cell.fill = upcoming_fill

            ws_detail.cell(row=detail_row, column=7, value=item.get("borrower_counsel", "N/A"))
            ws_detail.cell(row=detail_row, column=8, value=item.get("lender_counsel", "N/A"))
            ws_detail.cell(row=detail_row, column=9, value=item.get("source_filing", "N/A"))

            source = item.get("source_filing", "")
            filing_urls = result.get("filing_urls", {})
            url = filing_urls.get(source)

            source_cell = ws_detail.cell(row=detail_row, column=9, value=source if source else "N/A")
            if url:
                    source_cell.hyperlink = url
                    source_cell.font = Font(color="1F4E79", underline="single")

            detail_row += 1

    # Format detail sheet as Excel Table
    if detail_row > 2:
        table_ref = f"A1:{get_column_letter(len(detail_headers))}{detail_row - 1}"
        table = Table(displayName="DebtDetail", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws_detail.add_table(table)

    # Column widths for detail
    detail_widths = [10, 25, 25, 15, 15, 15, 25, 25]
    for col, width in enumerate(detail_widths, 1):
        ws_detail.column_dimensions[get_column_letter(col)].width = width

    wb.save(filename)
    print(f"\nSaved to {filename}")
# Note: The create_input_template function is separate and can be run independently to generate an input Excel file for users to fill in with company tickers and other details before running the analysis.
def create_input_template(filename="companies_input.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    headers = ["Ticker", "Company Name (reference only)", "Public/Private"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data
    sample = [
        ["AAPL", "Apple Inc.", "Public"],
        ["MSFT", "Microsoft", "Public"],
        ["GOOGL", "Alphabet", "Public"],
    ]
    
    for row, data in enumerate(sample, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    
    wb.save(filename)
    print(f"Input template saved to {filename}")

if not os.path.exists("companies_input.xlsx"):
    create_input_template()

# Main execution
def read_input(filename="companies_input.xlsx"):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    tickers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[0]
        is_private = str(row[2]).lower() == "private" if row[2] else False
        
        if not ticker:
            continue
        if is_private:
            print(f"Skipping {ticker} — private company, limited SEC data")
            continue
        tickers.append(ticker)
    
    return tickers

# Main execution
if __name__ == "__main__":
    tickers = read_input()
    results = []

    for ticker in tickers:
        try:
            result = analyze_company(ticker)
            results.append(result)
        except Exception as e:
            print(f"Skipping {ticker} — {str(e)}")

    if results:
        create_excel(results)
        print("Done!")
    else:
        print("No public companies to analyze")